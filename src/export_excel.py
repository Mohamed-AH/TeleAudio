"""
Phase 3 – Excel Export
Reads checkpoints/progress.json and writes two styled Excel files:
  output/full_archive.xlsx  — all records, 11 columns (updatedData format)
  output/khutba_only.xlsx   — khutba records only, 7 columns (khutba_archive format)
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
PROGRESS_PATH = ROOT / "checkpoints" / "progress.json"
OUTPUT_DIR = ROOT / "output"

# Styling constants (matching sample files)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
EVEN_FILL = PatternFill("solid", fgColor="D6E4F0")

FULL_HEADERS = [
    "S.No",
    "TelegramFileName",
    "Type",
    "SeriesName",
    "SequenceInSeries",
    "OriginalAuthor",
    "Location/Online",
    "Sheikh",
    "DateInGreg",
    "ClipLength",
    "Category",
]

KHUTBA_HEADERS = [
    "S.No",
    "TelegramFileName",
    "Type",
    "SeriesName",
    "Sheikh",
    "DateInGreg",
    "Category",
]

# Columns that contain Arabic and should be RTL/right-aligned
ARABIC_COLUMNS = {"SeriesName", "Sheikh", "OriginalAuthor", "Location/Online", "Type"}


def record_to_full_row(sno: int, rec: dict) -> list:
    ext = rec.get("extracted") or {}
    return [
        sno,
        rec.get("telegram_filename", ""),
        ext.get("Type", ""),
        ext.get("SeriesName", ""),
        ext.get("SequenceInSeries", ""),
        ext.get("OriginalAuthor", ""),
        ext.get("Location_Online", ""),
        ext.get("Sheikh", "حسن بن محمد منصور الدغريري"),
        ext.get("DateInGreg", rec.get("date_raw", "").replace(".", "/")),
        rec.get("clip_length", ""),
        ext.get("Category", ""),
    ]


def record_to_khutba_row(sno: int, rec: dict) -> list:
    ext = rec.get("extracted") or {}
    return [
        sno,
        rec.get("telegram_filename", ""),
        ext.get("Type", "Khutba"),
        ext.get("SeriesName", ""),
        ext.get("Sheikh", "حسن بن محمد منصور الدغريري"),
        ext.get("DateInGreg", rec.get("date_raw", "").replace(".", "/")),
        ext.get("Category", "Khutba"),
    ]


def apply_header(ws, headers: list[str]):
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_data_row(ws, row_idx: int, values: list, headers: list[str]):
    fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
    for col_idx, (value, header) in enumerate(zip(values, headers), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        if header in ARABIC_COLUMNS:
            cell.alignment = Alignment(horizontal="right", readingOrder=2)
        else:
            cell.alignment = Alignment(horizontal="left")


def auto_width(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    apply_header(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx, values in enumerate(rows, start=2):
        apply_data_row(ws, row_idx, values, headers)

    auto_width(ws, headers)
    wb.save(path)
    print(f"  Saved → {path.relative_to(ROOT)}  ({len(rows)} rows)")


def main():
    print("=== Phase 3: Excel Export ===\n")

    if not PROGRESS_PATH.exists():
        print("ERROR: checkpoints/progress.json not found. Run extract_metadata.py first.")
        sys.exit(1)

    with open(PROGRESS_PATH, encoding="utf-8") as f:
        data = json.load(f)

    done_records = [r for r in data["records"] if r["status"] == "done"]
    total = len(data["records"])
    print(f"Records: {len(done_records)} done / {total} total")

    if len(done_records) < total:
        pending = total - len(done_records)
        print(f"WARNING: {pending} records not yet extracted (status != done) — they will be skipped.")

    OUTPUT_DIR.mkdir(exist_ok=True)

    # Full archive
    full_rows = [record_to_full_row(i + 1, r) for i, r in enumerate(done_records)]
    write_workbook(
        OUTPUT_DIR / "full_archive.xlsx",
        "Full Archive",
        FULL_HEADERS,
        full_rows,
    )

    # Khutba only
    khutba_records = [r for r in done_records if (r.get("extracted") or {}).get("Type") == "Khutba"]
    khutba_rows = [record_to_khutba_row(i + 1, r) for i, r in enumerate(khutba_records)]
    write_workbook(
        OUTPUT_DIR / "khutba_only.xlsx",
        "Khutba Archive",
        KHUTBA_HEADERS,
        khutba_rows,
    )

    print(f"\nDone. Full archive: {len(full_rows)} rows | Khutba only: {len(khutba_rows)} rows")


if __name__ == "__main__":
    main()
