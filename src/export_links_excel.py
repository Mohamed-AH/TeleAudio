"""
Extract metadata from link-only messages and export to Excel.
Reads checkpoints/raw_link_messages.json and writes:
  output/links_archive.xlsx  (single sheet matching the full-archive format,
                               with primary link in the TelegramFileName column)
"""

import json
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT      = Path(__file__).parent.parent
RAW_PATH  = ROOT / "checkpoints" / "raw_link_messages.json"
OUT_PATH  = ROOT / "output" / "links_archive.xlsx"

# Reuse the shared extraction helpers
sys.path.insert(0, str(ROOT / "src"))
import extract_metadata as em

# ── styling ─────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")
EVEN_FILL    = PatternFill("solid", fgColor="D6E4F0")
ARABIC_COLS  = {"SeriesName", "Sheikh", "OriginalAuthor", "Location/Online", "Type"}

HEADERS = [
    "S.No",
    "TelegramFileName",   # holds the primary link URL
    "LinkType",           # YouTube / AudioURL / Other
    "Type",
    "SeriesName",
    "SequenceInSeries",
    "OriginalAuthor",
    "Location/Online",
    "Sheikh",
    "DateInGreg",
    "Category",
    "AllLinks",           # semicolon-separated full list for reference
]


# ── helpers ──────────────────────────────────────────────────────────────────

def strip_direction_marks(s: str) -> str:
    return re.sub(r"[​-‏‪-‮⁦-⁩﻿]", "", s)


def extract_record_from_link_msg(raw: dict) -> dict:
    """
    Build a pseudo-record compatible with em.extract_record by mapping
    link-message fields onto the audio-message schema.
    """
    # Synthesise an 'audio_title' from the message text (first meaningful line)
    text = raw.get("message_text", "")
    audio_title = ""
    for line in text.split("\n"):
        line_clean = re.sub(r"[#🔸🔹📌📕✏️🎙☑🔊◀️📥🔗🌀🎧◈]\S*", "", line)
        line_clean = re.sub(r"https?://\S+", "", line_clean).strip()
        line_clean = re.sub(r"^[\|\-–:\s]+|[\|\-–:\s]+$", "", line_clean).strip()
        if len(line_clean) > 6:
            audio_title = line_clean
            break

    pseudo = {
        "audio_title":      audio_title,
        "message_text":     text,
        "date_raw":         raw.get("date_raw", ""),
        "telegram_filename": raw.get("primary_link", ""),
        "clip_length":      "",
    }
    return em.extract_record(pseudo)


def make_row(sno: int, raw: dict, ext: dict) -> list:
    all_links_str = "; ".join(raw.get("all_links", []))
    return [
        sno,
        raw.get("primary_link", ""),
        raw.get("link_type", ""),
        ext.get("Type", ""),
        strip_direction_marks(ext.get("SeriesName", "")),
        ext.get("SequenceInSeries"),
        ext.get("OriginalAuthor"),
        ext.get("Location_Online", ""),
        ext.get("Sheikh", em.SHEIKH),
        ext.get("DateInGreg", raw.get("date_raw", "").replace(".", "/")),
        ext.get("Category", ""),
        all_links_str,
    ]


def write_workbook(path: Path, rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Links Archive"

    # Header row
    ws.append(HEADERS)
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for ri, values in enumerate(rows, 2):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, (v, h) in enumerate(zip(values, HEADERS), 1):
            cell = ws.cell(ri, ci, value=v)
            cell.fill = fill
            if h in ARABIC_COLS:
                cell.alignment = Alignment(horizontal="right", readingOrder=2)
            elif h in ("TelegramFileName", "AllLinks"):
                cell.alignment = Alignment(horizontal="left", wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal="left")

    # Auto-width columns
    for ci, h in enumerate(HEADERS, 1):
        col = get_column_letter(ci)
        max_len = len(h)
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, ci).value
            if val:
                # Don't let long URLs dictate column width
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col].width = min(max(max_len + 2, 12), 65)

    wb.save(path)


def main():
    print("=== Link Messages: Extract + Export ===\n")

    if not RAW_PATH.exists():
        print("ERROR: checkpoints/raw_link_messages.json not found. Run parse_links.py first.")
        sys.exit(1)

    with open(RAW_PATH, encoding="utf-8") as f:
        raw_records = json.load(f)

    print(f"Records loaded: {len(raw_records)}")

    rows = []
    from collections import Counter
    type_c = Counter()
    cat_c  = Counter()
    link_c = Counter()

    for i, raw in enumerate(raw_records):
        ext = extract_record_from_link_msg(raw)
        rows.append(make_row(i + 1, raw, ext))
        type_c[ext.get("Type", "")] += 1
        cat_c[ext.get("Category", "")] += 1
        link_c[raw.get("link_type", "")] += 1

    print(f"Link type:  {dict(link_c)}")
    print(f"Type:       {dict(type_c)}")
    print(f"Category:   {dict(cat_c)}")

    OUT_PATH.parent.mkdir(exist_ok=True)
    write_workbook(OUT_PATH, rows)
    print(f"\nSaved → {OUT_PATH.relative_to(ROOT)}  ({len(rows)} rows)")


if __name__ == "__main__":
    main()
